#import the necessary libraries 
import cv2 as cv 
#OpenCV is the huge open-source library for the computer vision, machine learning, and image 
# processing and now it plays a major role in real-time operation which is very important 
# in today's systems. By using it, one can process images and videos to identify objects, 
# faces, or even handwriting of a human.

import openpyxl 
      
# template12.png is the template of certificate 

template_path = 'template12.png'

# Excel file containing names of the participants 

details_path = 'name_list.xlsx'
 
# Output Path 

output_path = 'SL_assignment'
   
# Setting the font size and font colour in the format (B,G,R)

font_size = 3
font_size1 = 1.5
font_color = (255, 128, 0) 
font_color1 = (27, 115, 255)
font_color2 = (0,0,255)
font_color3 = (230,130,175)
   
# loading the details.xlsx workbook and grabbing the active sheet 

obj = openpyxl.load_workbook(details_path) 

sheet = obj.active 
  
# printing for the first 9 names in the excel sheet 

for i in range(2,11): 

    # grabs the row=i and column=1 cell  that contains the name value of that

    # cell-1 is stored in the variable certi_name 
    get_name = sheet.cell(row = i ,column = 1) 
    certi_name = get_name.value 
    
    # cell-2 is stored in the variable certi_position
    get_position = sheet.cell(row = i, column = 2)
    certi_position = get_position.value  
    
    #cell-3 is stored in the variable certi_event
    get_event = sheet.cell(row = i, column = 3)
    certi_event = get_event.value 

    #cell-4 is stored in the variable certi_clg
    get_clg = sheet.cell(row = i, column = 4)
    certi_clg = get_clg.value 
    
    # read the certificate template 

    img = cv.imread(template_path) 

    # choose the font from opencv 

    font = cv.FONT_HERSHEY_DUPLEX               
    font1 = cv.FONT_HERSHEY_TRIPLEX
    font2 = cv.FONT_HERSHEY_COMPLEX
    font3 = cv.FONT_HERSHEY_SIMPLEX

    # get the (x,y) coordinates where the name is to written on the template 

    cv.putText(img, certi_name, 
                (830,716),  
                font,  
                font_size, 
                font_color, 2) 

    cv.putText(img,certi_position,
                (1505,770),
                font1, 
                font_size1,
                font_color2, 3) 
    
    cv.putText(img,certi_event,
                (360,830),
                font2, 
                font_size1,
                font_color3, 4) 

    cv.putText(img,certi_clg,
                (450,770),
                font3, 
                font_size1,
                font_color1, 4) 

    # Output path along with the name of the certificate generated 
    # certi_path = output_path + 'Name of participant' + '.png'
    # Save the certificate                       

    cv.imwrite(f'SL_assignment/{certi_name}.jpg',img)